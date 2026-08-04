import requests
from bs4 import BeautifulSoup
import keyring
import re
from datetime import datetime, date
import sys
from requests import Response,exceptions
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os

def string_to_date(string: str) -> datetime.date | None: # type: ignore
    """
    Convert a date string in the format 'YYYY-MM-DD' to a date object.
    Returns None if the input is not a valid date string.
    """
    from datetime import datetime
    match = re.search(r"(\d{4}-\d{2}-\d{2})", string)
    birthdate_date = None
    if match:
        try:
            birthdate_date = datetime.strptime(match.group(1), "%Y-%m-%d").date()
            return birthdate_date
        
        except ValueError:
            birthdate_date = None

def string_to_name(strings: tuple[str, str]) -> str | None: # type: ignore
    """
    Convert a tuple of strings to a tuple of names (first name, last name).
    Returns None if the input is not valid.
    """
    first_name = ""
    last_name = ""
    for string in strings:
        if "fornavn" in string.lower():
            first_name = string.split(":")[1].replace("</li>", "").strip()
        elif "etternavn" in string.lower():
            last_name = string.split(":")[1].replace("</li>", "").strip()
    
    if first_name and last_name:
        return first_name + " " + last_name
    else:
        return None
    
def find_birthdate_line(page: str) -> tuple[datetime.date | None, str | None]: # type: ignore
    """
    Find the line in the page text that contains "født:" (case-insensitive).
    Returns the line if found, otherwise returns an empty string.
    """
    # Print the line(s) that contain "født:" (case-insensitive)
    for line in page.splitlines():
        if "fornavn" in line.lower():
            first_name = line
        if "etternavn" in line.lower():
            last_name = line
        if "født:" in line.lower():
            if "(krever innlogging)" in line.lower():
                print("You are not logged in, please log in with the authenticate script first.")
                sys.exit(1)
            return (string_to_date(line), string_to_name((first_name, last_name)))

def get_session(username: str) -> tuple[requests.Session, Response]:
    """
    Get the response from the page.
    Returns the response object.
    """

    login_url = "https://data.bowling.no/login/"
    session = requests.Session()

    # Step 1: Get the login page
    try:
        r = session.get(login_url, timeout=30)
    except requests.exceptions.Timeout:
        print("Request timed out. Please check your internet connection and try again.")
        sys.exit(1)

    soup = BeautifulSoup(r.text, "html.parser")

    loginid = soup.find("input", {"name": "loginid"})["value"] #type: ignore

    password = keyring.get_password(
        "NBF",
        username
    )

    # Step 2: Login
    payload = {
        "loginid": loginid,
        "UName": username,
        "UserPsw": password,
        "Submit": "log in"
    }

    r = session.post(login_url, data=payload)

    return session, r

def url_to_player_id(url: str) -> str | None:
    """
    Extract the player ID from the URL.
    Returns the player ID if found, otherwise returns None.
    """
    match = re.search(r"Id=(\d+)", url)
    if match:
        return match.group(1)
    else:
        return None

def get_player_id(licens_number: str, session: requests.Session | None = None) -> str:
    """
    Get the player ID from the page.
    Returns the player ID if found, otherwise returns None.
    """
    url = f"https://data.bowling.no/ajax/search/list.php?query={licens_number}"
    if session:
        response = session.get(url)
    else:
        response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    # Find the player ID in the page
    number = re.search(r'Id=(\d+)', soup.text).group(1) #type: ignore
    return number

def get_class(_date: datetime, year: int) -> str:
    birth = _date.date() if isinstance(_date, datetime) else _date

    # Calculate cutoffs dynamically
    junior_cutoff = date(year - 18, 9, 1)       # Sep 1, year-18 or later
    senior_cutoff = date(year - 18, 8, 31)      # Aug 31, year-18 or earlier
    u23_cutoff = date(year - 23, 7, 18)         # Jul 18, year-23 or later
    veteran_cutoff = date(year - 62, 7, 18)     # Jul 18, year-62 or earlier
    retired_cutoff = date(year - 67, 7, 18)     # Jul 18, year-67 or earlier

    # Exclusive priority (oldest groups first)
    if birth <= retired_cutoff:
        return "P"       # Retired
    elif birth <= veteran_cutoff:
        return "V"       # Veteran
    elif birth <= senior_cutoff:
        return ""       # Senior
    elif birth >= u23_cutoff and birth < junior_cutoff:
        return "U-23"       # U-23
    elif birth >= junior_cutoff:
        return "Jr"       # Junior
    

    return "Ukjent"

def get_participant_from_excel(file_path: str) -> list[list[str,str,str]]: # type: ignore
    """
    Get the participant list from the Excel file.
    Returns a list of tuples containing (license number, name).
    """
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    participants = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # type: ignore skip the header row
        license_number = str(row[0]).strip()  # Assuming license number is in the first column
        name = str(row[1]).strip()  # Assuming name is in the second column
        participants.append([license_number, name,""])  # Append an empty string for the class

    return participants

def save_participants_to_excel(participants: list[list[str,str,str]], file_path: str) -> None: # type: ignore

    # Create clean dataframe
    participants_df = pd.DataFrame(
        participants,
        columns=["Lisens", "Navn", "Klasse"]
    )

    participants_df = participants_df.sort_values(by="Klasse")
    
    file_path = get_unique_filename(file_path)
    participants_df.to_excel(file_path, index=False)

    excel_format(file_path)
    print(f"\nParticipants saved to {file_path}")

def get_unique_filename(filepath: str) -> str:
    """Generates a unique filename by appending a counter if the file already exists"""
    base, ext = os.path.splitext(filepath)
    candidate = filepath
    counter = 1

    if os.path.exists(candidate):
        option = input("A file already exist, do you want to overwrite it?(y/n): ")
        while option.lower() not in ["y","n"]:
            option = input("You have to write either y or n")

        if option =="n":
            while os.path.exists(candidate):
                candidate = f"{base}_({counter}){ext}"
                counter += 1

    return candidate

def excel_format(filepath: str): 
    """Formating the excel sheets with bold fonts, bottom borders and column width"""
    wb = load_workbook(f"{filepath}")
    ws = wb["Sheet1"]

    # Auto-fit column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column) #type: ignore

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2


    wb.save(f"{filepath}")

def get_classified_participants(file_path: str) -> list[list[str]]:
    df = pd.read_excel(file_path, sheet_name="Sheet1", header=None, dtype=object)
    df = df.fillna('')                      # replace NaN with empty string
    df = df.reindex(columns=[0, 1, 2], fill_value='')  # ensure at least 3 columns

    return df.iloc[:, :3].astype(str).values.tolist()

def get_new_participants(classified_list: list, unclassified_list: list) -> list:

    licensnumbers = {row[0] for row in classified_list}
    new_list = [row for row in unclassified_list if row[0] not in licensnumbers]

    return new_list

def update_options():
    print("\nDo you want to add only new participants or update all participants?: ")
    print("\n1. Add new Participants")
    print("2. Update existing Participants")
    option = input("\nEnter your option: ")

    while option not in ["1","2"]:
        option = input("Enter either 1 or 2")

    return option
